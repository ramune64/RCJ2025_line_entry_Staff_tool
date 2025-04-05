""" import glob

court_path = "templates/court/*.html"
court_files = glob.glob(court_path)
print(court_files)
for html_file in court_files:
    print(html_file)
    route_name = html_file[:-5].replace("templates/court\\","")
    print(route_name) """

""" def numatan():
    print("aaaaaaaaa said numatan") """
""" import win32com.client
import time

# Excelを起動
def open_excel():
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True  # Excelを画面に表示（Falseにすると非表示）
    print("Excelを起動しました。")
    return excel

# Excelを終了
def close_excel(excel):
    excel.Quit()
    print("Excelを閉じました。")

excel_app = open_excel()
try:
    input("Excelを閉じるにはEnterキーを押してください...")
finally:
    # Excelを閉じる
    close_excel(excel_app) """
#time.sleep(5)
#close_excel(excel_app)
""" import openpyxl

file_name = "result.xlsx"
sheet_name = 'Sheet1'

#wb = openpyxl.Workbook()#ワークブックの作成
wb = openpyxl.load_workbook(file_name,data_only=True)#ワークブックの読み込み
ws = wb[sheet_name]
cell1 = ws['A2']
cell1.value = "ddd"
print(cell1.value)


wb.save(file_name)
wb.close()

a = "eeaee"

a.split() """
""" for i in range(2,5):
    print(i) """
#open_excel()