from flask import Flask, jsonify ,render_template,request
import time
import glob
import openpyxl
from datetime import timedelta
import os
import shutil

excel_name = "result_4update.xlsx"
excel_name2 = "result.xlsx"
sheet_name = 'Sheet1'


app = Flask(__name__)

def create_text_file(content,scoring,tname,ch_num_panels_list,max,gimick_note_list):
    gimick_note = list(gimick_note_list)
    ch_num_panels = list(ch_num_panels_list)
    print("popints",ch_num_panels)
    # テキストファイルを作成してデータを書き込む
    template_file = open("static/template2.txt","r",encoding='utf-8')
    data = template_file.read()
    template_file.close()
    #print(data)
    con = str(content)
    con2 = '<div class="scoring" id="scoring">'+str(scoring)+'</div>'
    con2 = con2.replace("<tbody>","")
    con2 = con2.replace("</tbody>","")
    #con = con.replace("tbody","table")
    for p in ch_num_panels:
        data += '<div class="panels_num">'+str(p)+'</div>'
    for q in gimick_note:
        data += '<div class="gimick_note">'+str(q)+'</div>'
    data += '<div id="max_point">'+str(max)+'</div>'
    con = con.replace("<table>",'<table class="field">')
    con = con.replace("<tbody>","")
    con = con.replace("</tbody>","")
    con = data + con
    new_file_name = "templates/court/"+str(tname)+".html"
    print(new_file_name)
    with open(new_file_name, mode='w',encoding='utf-8') as f:
        print()
        f.write(con+'</table><p class="margin_top_20px"><b>チーム名</b></p><form action="#" method="post" id="form3"><input type="text" id="team" name="team" size="40" required /><br><label for="">チーム番号:</label><input type="number" max="10" min="1" id="num_times" required /><br><input type="submit" id="st_scoreing" value="採点画面へ"></form></div>'+con2+'<script src="https://cdnjs.cloudflare.com/ajax/libs/jspdf/2.5.2/jspdf.umd.min.js"></script><script src="/static/app.js"></script><script src="/static/app2.js"></script></body></html>')
        f.close()
    
    #print(con)
    print("Text file created1!")
    return new_file_name.replace("templates/","")
    #webview.create_window("RCJ Rescue Line Entry 2025 Field Maker", new_file_name, js_api=API())

def create_scoring_file(field_table,score_table,tname,ch_num_panels_list,max,gimick_note_list,num_times,num_times_2):#,field_table,score_table,name
    gimick_note = list(gimick_note_list)
    ch_num_panels = list(ch_num_panels_list)
    print("popints",ch_num_panels)
    print("clicked")
    template_file = open("static/template2.txt","r",encoding='utf-8')
    data = template_file.read()
    template_file.close()

    for p in ch_num_panels:
        data += '<div class="panels_num">'+str(p)+'</div>'
    for q in gimick_note:
        data += '<div class="gimick_note">'+str(q)+'</div>'
    data += '<div id="max_point2">'+str(max)+'</div>'
    data += '<div id="num_time_wa">'+num_times+'</div>'
    data += '<div id="num_time_2_wa">'+num_times_2+'</div>'
    con = str(field_table)
    con = con.replace("<table>",'<table class="field">')
    con = data + con + '</table></div>'

    con2 = str(score_table)
    con2 = con2.replace("<table>",' id="scoring_table" border="1"')
    con2 = con2.replace("<tbody>","")
    con2 = con2.replace("</tbody>","")
    con2 = '<div class="scoring" id="scoring">'+con2
    new_file_name = "templates/scoring/"+"scoring_"+str(tname)+"_"+str(num_times)+"_"+str(num_times_2)+".html"
    point = '<h2>得点:<b style="font-size: 25pt;" id="current_score">'+str(0)+'</b>/'+str(max)+'点</h2>'
    others = '<h2>競技時間</h2><form action="#" id="form5"><input type="number" min="0" max="8" class="last_decision" id="min" required><label>分</label><input type="number" min="0" max="59" id="sec" class="last_decision" required><label>秒</label><br><input type="submit" value="この内容で決定" class="last_decision last_button"></form><h2 id="team_name">チーム名:'+str(name)+'</h2><div class="copy_flex"><h2 class="sign">主審署名</h2><canvas id="drawingCanvas2" class="drawingCanvas" width="900" height="300"></canvas><button id="clearButton2" class="nocopy">リセット</button><h2 class="sign">副審署名</h2><canvas id="drawingCanvas3" class="drawingCanvas" width="900" height="300"></canvas><button id="clearButton3" class="nocopy">リセット</button><h2 class="sign">キャプテン署名</h2><canvas id="drawingCanvas" class="drawingCanvas" width="900" height="300"></canvas><button id="clearButton" class="nocopy">リセット</button></div><form action="#" id="form6" class="nocopy noprint" style="display: none;" ><input type="submit" value="印刷" id="copy" class="noprint"></form>'
    with open(new_file_name, 'w',encoding='utf-8') as f:
        f.write(con+" \n"+con2+'\n</table>\n'+point+'</div>\n'+others+'<script src="https://cdnjs.cloudflare.com/ajax/libs/jspdf/2.5.2/jspdf.umd.min.js"></script><script src="/static/app.js"></script><script src="/static/app3.js"></script>\n</body>\n</html>')
        f.close()
    return new_file_name.replace("templates/","")
    #webview.create_window("RCJ Rescue Line Entry 2025 scoring "+ str(name), new_file_name, js_api=API())

def create_result_file(content,team):
    team = str(team).replace("チーム名:","")
    new_file_name = "templates/result/result_"+str(team)+"_"+str(2)+".html"
    with open(new_file_name,mode="w",encoding='utf-8') as f:
        f.write(content)
        f.close()

def write_excel(score,play_time,teamnum,teamname):
    wb = openpyxl.load_workbook(excel_name,data_only=True)#ワークブックの読み込み
    ws = wb.active
    cell_num = 2
    max_cell = 2
    while True:
        cell = "A"+str(cell_num)
        if ws[cell].value == teamnum[:-3]:
            break
        if ws[cell].value == None:
            break
        cell_num += 1
    while True:
        cell = "A"+str(max_cell)
        if ws[cell].value == None:
            break
        max_cell += 1
    min,other = str(play_time).split("分")
    sec = other.split("秒")[0]
    if int(teamnum[-3]) == 1:
        ws["A"+str(cell_num)].value = teamnum[:-3]
        ws["B"+str(cell_num)].value = teamname.replace("チーム名:","")
        ws["C"+str(cell_num)].value = int(score)
        ws["F"+str(cell_num)].value = timedelta(minutes=int(min), seconds=int(sec))
        ws["F"+str(cell_num)].number_format = 'mm分ss秒'  # 時間フォーマット
        formula1 = "=SUM(C{}:D{})".format(cell_num,cell_num)
        formula2 = "=SUM(F{}:G{})".format(cell_num,cell_num)
        formula3 = "=E{}*100000-(MINUTE(H{})*60+SECOND(H{}))".format(cell_num,cell_num,cell_num)
        formula4 = "=RANK(J{},$J$2:$J$100)".format(cell_num)
        print(formula1)
        print(formula2)
        for k in range(2,max_cell+1):
            formula1 = "=SUM(C{}:D{})".format(k,k)
            formula2 = "=SUM(F{}:G{})".format(k,k)
            formula3 = "=E{}*100000-(MINUTE(H{})*60+SECOND(H{}))".format(k,k,k)
            formula4 = "=RANK(J{},$J$2:$J$100)".format(k)
            print(k)
            ws["E"+str(k)].value = formula1
            ws["H"+str(k)].value = formula2
            ws["H"+str(k)].number_format = 'mm分ss秒'
            ws["J"+str(k)].value = formula3
            ws["I"+str(k)].value = formula4
    else:
        ws["D"+str(cell_num)].value = int(score)
        ws["G"+str(cell_num)].value = timedelta(minutes=int(min), seconds=int(sec))
        ws["G"+str(cell_num)].number_format = 'mm分ss秒'  # 時間フォーマット
        formula1 = "=SUM(C{}:D{})".format(cell_num,cell_num)
        formula2 = "=SUM(F{}:G{})".format(cell_num,cell_num)
        formula3 = "=E{}*100000-(MINUTE(H{})*60+SECOND(H{}))".format(cell_num,cell_num,cell_num)
        formula4 = "=RANK(J{},$J$2:$J$100)".format(cell_num)
        print(formula1)
        print(formula2)
        for k in range(2,max_cell):
            formula1 = "=SUM(C{}:D{})".format(k,k)
            formula2 = "=SUM(F{}:G{})".format(k,k)
            formula3 = "=E{}*100000-(MINUTE(H{})*60+SECOND(H{}))".format(k,k,k)
            formula4 = "=RANK(J{},$J$2:$J$100)".format(k)
            print(k)
            ws["E"+str(k)].value = formula1
            ws["H"+str(k)].value = formula2
            ws["H"+str(k)].number_format = 'mm分ss秒'
            ws["J"+str(k)].value = formula3
            ws["I"+str(k)].value = formula4
    
    wb.save(excel_name)
    wb.close()
    shutil.copy(excel_name,excel_name2)


path = ""
path2 = ""


@app.route('/')
def home():
    # templates/index.htmlをレンダリングして返す
    return render_template('index.html')

@app.route('/newcourt')
def newcourt():
    time.sleep(0.5)
    global path
    # templates/index.htmlをレンダリングして返す
    print(path)
    return render_template(path)

@app.route('/scoring')
def scoring():
    print("scor2")
    time.sleep(0.5)
    global path2
    # templates/index.htmlをレンダリングして返す
    print(path2)
    return render_template(path2)


@app.route('/create_text_file', methods=['POST'])
def call_create_text_file():
    data = request.get_json()
    content=data.get("content",0)
    score_table=data.get("score_table",0)
    teamname=data.get("team",0)
    ch_num_panels_list=data.get("ch_num_panels_list",0)
    max_point=data.get("max",0)
    gimick_note_list=data.get("gimick_note_list",0)
    global path
    path = create_text_file(content,score_table,teamname,ch_num_panels_list,max_point,gimick_note_list)
    print(path)
    #global path
    return "File created successfully", 200

@app.route('/api/create_scoring_file', methods=['POST'])
def call_create_scoring_file():
    data = request.get_json()
    field_table=data.get("field_table",0)
    score_table=data.get("score_table",0)
    global name
    name=data.get("team",0)
    ch_num_panels_list=data.get("ch_num_panels_list",0)
    max_point=data.get("max",0)
    gimick_note_list=data.get("gimick_note_list",0)
    num_times = data.get("num_times",0)
    num_times_2 = data.get("num_times_2",0)
    global path2
    path2 = create_scoring_file(field_table,score_table,name,ch_num_panels_list,max_point,gimick_note_list,num_times,num_times_2)
    return "File created successfully", 200

@app.route("/api/create_result_file", methods=['POST'])
def call_create_result_file():
    data = request.get_json()
    content=data.get("content",0)
    team = data.get("team",0)
    score = data.get("score",0)
    play_time = data.get("play_time",0)
    team_num = data.get("team_num",0)
    create_result_file(content,team)
    write_excel(score,play_time,team_num,team)
    return "File created successfully", 200

@app.route("/api/create_court_list", methods=['POST'])
def call_create_court_list():
    court_path = "templates/court/*.html"
    court_files = glob.glob(court_path)
    text_content = ""
    for html_file in court_files:
        route_name = html_file[:-5].replace("templates/court\\","")
        text_content += f'<a href="{route_name}" target="_blank"><p style="font-size: 20pt;">{route_name}</p></a>\n'
    #print(text_content)
    return jsonify({'result':text_content})


court_path = "templates/court/*.html"
court_files = glob.glob(court_path)
for html_file in court_files:
    #route_name = html_file[:-5].replace("templates/court\\","")
    route_name = os.path.splitext(os.path.basename(html_file))[0]  # ファイル名から拡張子を除く
    print(route_name)
    # 動的に関数を生成
    def create_render_function(route):
        def render_function():
            return render_template(f"court/{route}.html")
        return render_function
    """ def render_court_page():
        return render_template(f"/court/{route_name}.html") """
    # ルートを登録
    app.add_url_rule(f"/{route_name}", route_name, create_render_function(route_name))



if __name__ == '__main__':
    app.run(debug=True,host='0.0.0.0',port=5000)